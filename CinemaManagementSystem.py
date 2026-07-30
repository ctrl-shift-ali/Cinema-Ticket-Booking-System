from flask import Flask, render_template, request, redirect, url_for, send_file, abort
from pathlib import Path
from datetime import datetime, timedelta
import uuid
import random
from openpyxl import Workbook, load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas as pdfcanvas
from reportlab.pdfbase.pdfmetrics import stringWidth

app = Flask(__name__, template_folder='templates', static_folder='static')
app.config['TEMPLATES_AUTO_RELOAD'] = True
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0

# Environment-aware file paths for serverless deployment
import os
is_vercel = os.environ.get('VERCEL') == '1'

baseDir = Path(__file__).resolve().parent
if is_vercel:
    # Use /tmp for Vercel serverless (writable directory)
    dataDir = Path('/tmp/data')
    ticketDir = Path('/tmp/tickets')
else:
    # Use local directories for development
    dataDir = baseDir / "data"
    ticketDir = baseDir / "tickets"

# Create directories if they don't exist
try:
    dataDir.mkdir(exist_ok=True, parents=True)
    ticketDir.mkdir(exist_ok=True, parents=True)
except Exception as e:
    print(f"Warning: Could not create directories: {e}")

excelPath = dataDir / "bookings.xlsx"

#For Adding Seats
roomCapacity = 70
seatsPerRow = 10
rowLetters = ["A", "B", "C", "D", "E", "F", "G"]

movieOptions = [
    {
        "title": "Neon Horizon",
        "description": "A bright cyber adventure with a futuristic soundtrack.",
        "roomNumber": 1,
        "slots": [
            {"showTime": "17:00", "price": 1200},
            {"showTime": "20:00", "price": 1600},
            {"showTime": "23:00", "price": 1800},
        ],
    },
    {
        "title": "Midnight Echo",
        "description": "A suspenseful night thriller with an unforgettable finale.",
        "roomNumber": 2,
        "slots": [
            {"showTime": "18:30", "price": 1400},
            {"showTime": "21:30", "price": 1700},
        ],
    },
    {
        "title": "Golden Hour",
        "description": "A warm family drama with rich visuals and heartfelt storytelling.",
        "roomNumber": 3,
        "slots": [
            {"showTime": "16:00", "price": 1100},
            {"showTime": "19:00", "price": 1500},
            {"showTime": "22:00", "price": 1700},
        ],
    },
    {
        "title": "Insidious: Out Of The Further",
        "description": "Gemma, a young mother accidentally pulls Malevolent Entities from The Further into the real world when she travels there.",
        "roomNumber": 4,
        "slots": [
            {"showTime": "16:00", "price": 1100},
            {"showTime": "20:00", "price": 1500},
            {"showTime": "02:00", "price": 1700},
                ],
    },
    {
        "title": "Interstellar",
        "description": "Science Fiction (Sci-Fi).",
        "roomNumber": 5,
        "slots": [
            {"showTime": "13:00", "price": 1100},
            {"showTime": "18:00", "price": 1500},
            {"showTime": "21:00", "price": 1700},
        ],
    },
]

expectedColumns = [
    "bookingId",
    "personName",
    "ticketCount",
    "seatNumbers",
    "movieTitle",
    "showTime",
    "departTime",
    "roomNumber",
    "showDate",
    "pricePerSeat",
]

def seatLabelFromNumber(seatNumber):
    rowIndex = (seatNumber - 1) // seatsPerRow
    colIndex = (seatNumber - 1) % seatsPerRow + 1
    return f"{rowLetters[rowIndex]}{colIndex}"


def findMovieByTitle(movieTitle):
    return next((movie for movie in movieOptions if movie["title"] == movieTitle), None)


def findSlot(movie, showTime):
    if not movie:
        return None
    return next((slot for slot in movie["slots"] if slot["showTime"] == showTime), None)


def getPermanentBookedSeats(roomNumber):
    rnd = random.Random(f"room-{roomNumber}")
    return set(rnd.sample(range(1, roomCapacity + 1), 3))


def getBookedSeatsForRoom(roomNumber, bookings):
    bookedSeats = set(getPermanentBookedSeats(roomNumber))
    for row in bookings:
        if str(row.get("roomNumber", "")) == str(roomNumber):
            seatNumbers = row.get("seatNumbers", "")
            if isinstance(seatNumbers, str) and seatNumbers:
                for seat in seatNumbers.split(","):
                    if seat.strip():
                        bookedSeats.add(int(seat.strip()))
    return bookedSeats


def getRoomStatusInfo(roomNumber, bookings):
    bookedSeats = getBookedSeatsForRoom(roomNumber, bookings)
    return {"available": roomCapacity - len(bookedSeats), "booked": len(bookedSeats)}


def buildSeatGrid(roomNumber, bookings):
    bookedSeats = set(getBookedSeatsForRoom(roomNumber, bookings))

    grid = []
    for rowIndex, rowLetter in enumerate(rowLetters):
        rowSeats = []
        for col in range(1, seatsPerRow + 1):
            seatNumber = rowIndex * seatsPerRow + col
            rowSeats.append(
                {
                    "number": seatNumber,
                    "label": seatLabelFromNumber(seatNumber),
                    "booked": seatNumber in bookedSeats,
                }
            )
        grid.append({"rowLetter": rowLetter, "seats": rowSeats})
    return grid

def createExcelFile():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "bookings"
    sheet.append(expectedColumns)
    workbook.save(excelPath)
    workbook.close()


def loadBookingsData():
    if not excelPath.exists():
        createExcelFile()
    workbook = load_workbook(excelPath, data_only=True)
    sheet = workbook["bookings"]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows or rows[0] != tuple(expectedColumns):
        workbook.close()
        createExcelFile()
        workbook = load_workbook(excelPath, data_only=True)
        sheet = workbook["bookings"]
        rows = list(sheet.iter_rows(values_only=True))
    workbook.close()
    if not rows:
        return []
    return [
        dict(zip(expectedColumns, row))
        for row in rows[1:]
        if any(cell is not None and str(cell).strip() != "" for cell in row)
    ]


def saveBookingsData(bookings):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "bookings"
    sheet.append(expectedColumns)
    for booking in bookings:
        sheet.append([booking.get(column, "") for column in expectedColumns])
    workbook.save(excelPath)
    workbook.close()


def buildDepartTimeValue(showTime):
    showDateTime = datetime.strptime(showTime, "%H:%M")
    departDateTime = showDateTime + timedelta(hours=3)
    return departDateTime.strftime("%H:%M")

@app.route("/")
def indexPage():
    bookings = loadBookingsData()
    movieData = []
    for movie in movieOptions:
        roomStatus = getRoomStatusInfo(movie["roomNumber"], bookings)
        movieData.append(
            {
                "title": movie["title"],
                "description": movie["description"],
                "roomNumber": movie["roomNumber"],
                "availableSeats": roomStatus["available"],
                "bookedSeats": roomStatus["booked"],
                "slots": movie["slots"],
            }
        )
    return render_template(
        "index.html",
        movies=movieData,
        status=request.args.get("status"),
        bookingId=request.args.get("bookingId"),
        error=request.args.get("error"),
    )


@app.route("/seats")
def seatsPage():
    movieTitle = request.args.get("movieTitle", "").strip()
    showTime = request.args.get("showTime", "").strip()
    personName = request.args.get("personName", "").strip()
    try:
        ticketCount = int(request.args.get("ticketCount", "0"))
    except ValueError:
        ticketCount = 0

    movie = findMovieByTitle(movieTitle)
    slot = findSlot(movie, showTime)

    if not movie or not slot or ticketCount <= 0 or not personName:
        return redirect(url_for("indexPage", error="Please choose a movie, a showtime and enter your details first."))

    bookings = loadBookingsData()
    seatGrid = buildSeatGrid(movie["roomNumber"], bookings)
    availableCount = roomCapacity - len(getBookedSeatsForRoom(movie["roomNumber"], bookings))

    return render_template(
        "seats.html",
        movieTitle=movie["title"],
        roomNumber=movie["roomNumber"],
        showTime=showTime,
        price=slot["price"],
        personName=personName,
        ticketCount=ticketCount,
        seatGrid=seatGrid,
        availableCount=availableCount,
        error=request.args.get("error"),
    )


@app.route("/book", methods=["POST"])
def bookTicketPage():
    personName = request.form.get("personName", "").strip()
    movieTitle = request.form.get("selectedMovieTitle", "").strip()
    showTime = request.form.get("selectedShowTime", "").strip()
    try:
        ticketCount = int(request.form.get("ticketCount", "0"))
    except ValueError:
        ticketCount = 0
    seatNumbersRaw = request.form.get("seatNumbers", "").strip()

    movie = findMovieByTitle(movieTitle)
    slot = findSlot(movie, showTime)

    if not movie or not slot or ticketCount <= 0 or not personName or not seatNumbersRaw:
        return redirect(url_for("indexPage", error="Booking details were incomplete. Please try again."))

    try:
        requestedSeats = sorted({int(seat) for seat in seatNumbersRaw.split(",") if seat.strip()})
    except ValueError:
        return redirect(url_for("indexPage", error="Invalid seat selection."))

    if len(requestedSeats) != ticketCount:
        return redirect(url_for(
            "seatsPage", movieTitle=movieTitle, showTime=showTime,
            personName=personName, ticketCount=ticketCount,
            error=f"Please select exactly {ticketCount} seat(s)."
        ))

    bookings = loadBookingsData()
    alreadyBooked = getBookedSeatsForRoom(movie["roomNumber"], bookings)
    conflict = [seat for seat in requestedSeats if seat in alreadyBooked]
    if conflict:
        conflictLabels = ", ".join(seatLabelFromNumber(s) for s in conflict)
        return redirect(url_for(
            "seatsPage", movieTitle=movieTitle, showTime=showTime,
            personName=personName, ticketCount=ticketCount,
            error=f"Seat(s) {conflictLabels} were just booked by someone else. Please pick again."
        ))

    departTime = buildDepartTimeValue(showTime)
    bookingId = uuid.uuid4().hex[:8]
    showDate = datetime.now().strftime("%d/%m/%y")

    newRow = {
        "bookingId": bookingId,
        "personName": personName,
        "ticketCount": ticketCount,
        "seatNumbers": ",".join(str(seat) for seat in requestedSeats),
        "movieTitle": movie["title"],
        "showTime": showTime,
        "departTime": departTime,
        "roomNumber": movie["roomNumber"],
        "showDate": showDate,
        "pricePerSeat": slot["price"],
    }

    bookings.append(newRow)
    saveBookingsData(bookings)
    createTicketPdfFile(newRow, requestedSeats)
    return redirect(url_for("indexPage", status="success", bookingId=bookingId))

@app.route("/download/<bookingId>")
def downloadTicketPage(bookingId):
    ticketPath = ticketDir / f"{bookingId}.pdf"
    if not ticketPath.exists():
        abort(404)
    return send_file(ticketPath, as_attachment=True, download_name=f"{bookingId}.pdf")

ACCENT = colors.HexColor("#7c3aed")
ACCENT_DARK = colors.HexColor("#4c1d95")
CARD_BG = colors.HexColor("#c0264f")
CARD_BG_LIGHT = colors.HexColor("#d1416a")
WHITE = colors.white

def drawPerforationLine(c, x, yTop, yBottom):
    c.saveState()
    c.setDash(1, 3)
    c.setLineWidth(1)
    c.setStrokeColor(colors.HexColor("#f4c2d3"))
    c.line(x, yBottom, x, yTop)
    c.restoreState()

def drawBarcode(c, x, y, width, height, seed):
    c.saveState()
    c.setFillColor(colors.HexColor("#1f2937"))
    barX = x
    import random
    rnd = random.Random(seed)
    while barX < x + width:
        barW = rnd.choice([1.2, 1.8, 2.4, 3.0])
        c.rect(barX, y, barW, height, fill=1, stroke=0)
        barX += barW + rnd.choice([1.5, 2.2, 3.0])
    c.restoreState()

def drawTicketCard(c, originX, originY, width, height, booking, seatLabel, seatIndex):
    stubWidth = width * 0.34
    mainWidth = width - stubWidth

    # Card backgrounds
    c.setFillColor(CARD_BG)
    c.roundRect(originX, originY, mainWidth, height, 5, fill=1, stroke=0)
    c.setFillColor(CARD_BG_LIGHT)
    c.roundRect(originX + mainWidth, originY, stubWidth, height, 5, fill=1, stroke=0)
    # square off the shared edge so it reads as one card
    c.setFillColor(CARD_BG)
    c.rect(originX + mainWidth - 8, originY, 8, height, fill=1, stroke=0)
    c.setFillColor(CARD_BG_LIGHT)
    c.rect(originX + mainWidth, originY, 8, height, fill=1, stroke=0)

    drawPerforationLine(c, originX + mainWidth, originY + height, originY)

    top = originY + height
    textX = originX + 6 * mm
    rightEdge = originX + mainWidth - 6 * mm

    # ---- Header bar ----
    headerH = 6 * mm
    c.setFillColor(colors.HexColor("#2d2140"))
    c.rect(originX, top - headerH, mainWidth, headerH, fill=1, stroke=0)
    c.setFillColor(WHITE)
    c.setFont("Helvetica-Bold", 7)
    c.drawString(textX, top - headerH + 1.8 * mm, "TICKET")
    c.drawRightString(rightEdge, top - headerH + 1.8 * mm, "\u2605 CINEMA TICKET \u2605")

    # ---- Movie title (truncated with ellipsis if too wide) ----
    titleFontSize = 10
    maxTitleWidth = rightEdge - textX
    title = booking["movieTitle"]
    if stringWidth(title, "Helvetica-Bold", titleFontSize) > maxTitleWidth:
        while title and stringWidth(title + "\u2026", "Helvetica-Bold", titleFontSize) > maxTitleWidth:
            title = title[:-1]
        title = title.rstrip() + "\u2026"
    c.setFillColor(WHITE)
    c.setFont("Helvetica-Bold", titleFontSize)
    c.drawString(textX, top - 11.5 * mm, title)

    # ---- Field row (single row, six compact columns) ----
    labelY = top - 17 * mm
    valueY = top - 20.5 * mm
    colWidth = (mainWidth - 12 * mm) / 6

    def field(colIndex, labelText, valueText):
        x = textX + colIndex * colWidth
        c.setFont("Helvetica", 5.2)
        c.setFillColor(colors.HexColor("#f4c2d3"))
        c.drawString(x, labelY, labelText)
        c.setFont("Helvetica-Bold", 7)
        c.setFillColor(WHITE)
        c.drawString(x, valueY, valueText)

    fields = [
        ("ROOM", str(booking["roomNumber"])),
        ("SEAT", seatLabel),
        ("PRICE", f"PKR {booking['pricePerSeat']}"),
        ("DATE", booking["showDate"]),
        ("SHOW", booking["showTime"]),
        ("DEPART", booking["departTime"]),
    ]
    for colIndex, (labelText, valueText) in enumerate(fields):
        field(colIndex, labelText, valueText)

    # ---- Footer (booking + ticket reference) ----
    c.setFont("Helvetica", 6)
    c.setFillColor(colors.HexColor("#f4c2d3"))
    c.drawString(textX, top - 29.5 * mm, f"NO. {booking['bookingId'].upper()}{seatIndex:02d}")

    # ---- Stub (seat number side) — barcode lives here only ----
    stubX = originX + mainWidth
    stubCenter = stubX + stubWidth / 2

    c.setFillColor(WHITE)
    c.setFont("Helvetica-Bold", 20)
    c.drawCentredString(stubCenter, top - 10 * mm, seatLabel)

    c.setFont("Helvetica", 6.5)
    c.setFillColor(colors.HexColor("#f4c2d3"))
    c.drawCentredString(stubCenter, top - 15 * mm, "CINEMA TICKET")

    c.setFont("Helvetica-Bold", 7)
    c.setFillColor(WHITE)
    c.drawCentredString(stubCenter, top - 21 * mm, "ROOM " + str(booking["roomNumber"]))
    c.setFont("Helvetica", 7)
    c.drawCentredString(stubCenter, top - 26 * mm, booking["showTime"])

    drawBarcode(c, stubX + 5 * mm, top - 32 * mm, stubWidth - 10 * mm, 3 * mm, seatLabel + booking["bookingId"])


def createTicketPdfFile(booking, seatNumbers):
    ticketPath = ticketDir / f"{booking['bookingId']}.pdf"
    pageW, pageH = letter
    c = pdfcanvas.Canvas(str(ticketPath), pagesize=letter)

    marginX = 20 * mm
    marginTop = 10 * mm
    cardWidth = pageW - 2 * marginX
    cardHeight = 33 * mm
    cardGap = 4 * mm
    ticketsPerPage = 7
    originX = marginX

    for index, seatNumber in enumerate(seatNumbers, start=1):
        positionOnPage = (index - 1) % ticketsPerPage
        if positionOnPage == 0 and index != 1:
            c.showPage()
        topY = pageH - marginTop - positionOnPage * (cardHeight + cardGap)
        originY = topY - cardHeight
        seatLabel = seatLabelFromNumber(seatNumber)
        drawTicketCard(c, originX, originY, cardWidth, cardHeight, booking, seatLabel, index)

    c.save()


if __name__ == "__main__":
    app.run(debug=True)